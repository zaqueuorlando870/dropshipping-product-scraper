import time
import random
import requests
import pandas as pd
from io import BytesIO
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image as PILImage

# Setup Selenium
print("🚀 Starting Selenium WebDriver...")
service = Service("/Users/orlando/Desktop/aliexpress-scraper/chromedriver")  # <- Replace with your ChromeDriver path
options = webdriver.ChromeOptions()
options.add_argument("--headless=new")
driver = webdriver.Chrome(service=service, options=options)

# AliExpress category page
url = "https://computermania.co.za/collections/software/products/bitdefender-mobile-security-1-year-3-devices"
print(f"🌐 Loading category page: {url}")
driver.get(url)
time.sleep(5)
print("⏳ Waiting for page to load and initialize...")

product_links = [url]  # Just one product to process

# Universal helper functions
def get_element_text(driver, selectors):
    for by, selector in selectors:
        try:
            el = driver.find_element(by, selector)
            text = el.text.strip()
            if text:
                return text
        except:
            continue
    return ""

def get_image_url(driver):
    # Get main image (first found) and all images from #Product-Thumbnails
    # Main image logic (first found)
    selectors = [
        '#Product-Slider',
        "img[src*='jpg']",
        "img[src*='png']",
        "img.product-image",
        "img.main-image",
        "img.primary-image",
        "[itemprop='image']",
    ]

    all_images_set = set()
    first_image = None

    for sel in selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in elements:
                src = el.get_attribute('src')
                if src and ('jpg' in src or 'png' in src):
                    if not first_image:
                        first_image = src
                    all_images_set.add(src)
        except:
            continue

    # Now get all images from #Product-Thumbnails
    thumb_images = []
    try:
        thumb_container = driver.find_element(By.CSS_SELECTOR, '#Product-Thumbnails')
        thumb_els = thumb_container.find_elements(By.TAG_NAME, 'img')
        for el in thumb_els:
            src = el.get_attribute('src')
            if src and ('jpg' in src or 'png' in src):
                thumb_images.append(src)
    except Exception:
        pass

    all_images_list = list(all_images_set)
    all_images_str = ", ".join(all_images_list)
    thumb_images_str = ", ".join(thumb_images) if thumb_images else None

    return first_image, all_images_str if all_images_list else None, thumb_images_str

# Define common selectors
name_selectors = [
    (By.CSS_SELECTOR, '#root > div > div.pdp-body.pdp-wrap > div > div.pdp-body-top-left > div.pdp-info > div.pdp-info-right > div.title--wrap--UUHae_g > h1'),
    (By.CSS_SELECTOR, 'h1.product-title'),
    (By.CSS_SELECTOR, 'h1.title'),
    (By.CSS_SELECTOR, 'h1'),
    (By.CSS_SELECTOR, '[itemprop="name"]'),
    (By.CSS_SELECTOR, 'h2.product-name'),
]

price_selectors = [
    (By.CSS_SELECTOR, 'div.product-price-value'),
    (By.CSS_SELECTOR, '.price'),
    (By.CSS_SELECTOR, '[itemprop="price"]'),
    (By.CSS_SELECTOR, '.product-price'),
    (By.CSS_SELECTOR, '.price-current'),
]

desc_selectors = [
    (By.CSS_SELECTOR, '#ProductInfo-template--24701812048172__main-product > div:nth-child(13)'),
    (By.CSS_SELECTOR, '#description'),
    (By.CSS_SELECTOR, '.product-description'),
    (By.CSS_SELECTOR, '[itemprop="description"]'),
]

# Prepare Data
data = []

for idx, link in enumerate(product_links, start=1):
    print(f"\n📦 Scraping product {idx}/{len(product_links)}: {link}")
    driver.get(link)
    sleep_time = random.uniform(3, 5)
    print(f"⏳ Waiting {sleep_time:.2f}s for product page to load...")
    time.sleep(sleep_time)

    try:
        name = get_element_text(driver, name_selectors)
        price = get_element_text(driver, price_selectors)
        desc = get_element_text(driver, desc_selectors)
        image_url, all_images, thumb_images = get_image_url(driver)
        print(f"🔹 Product Name: {name[:60]}{'...' if len(name) > 60 else ''}")
        print(f"🔹 Price: {price}")
        print(f"🔹 Description length: {len(desc)} characters")
        print(f"🔹 Image URL: {image_url if image_url else 'None found'}")

        # Fetch category, brand, color, size if available
        specs = driver.find_elements(By.CSS_SELECTOR, 'ul.product-specs-list li')
        category = brand = color = size = ""
        if specs:
            print(f"🔍 Found product specs:")
        for spec in specs:
            text = spec.text.lower()
            if "category" in text:
                category = spec.text.split(":")[-1].strip()
                print(f"   - Category: {category}")
            elif "brand" in text:
                brand = spec.text.split(":")[-1].strip()
                print(f"   - Brand: {brand}")
            elif "color" in text:
                color = spec.text.split(":")[-1].strip()
                print(f"   - Color: {color}")
            elif "size" in text:
                size = spec.text.split(":")[-1].strip()
                print(f"   - Size: {size}")

        # 'Image' is the first image link (main image), 'Images' is all thumbnails as comma-separated string
        images_col = thumb_images if isinstance(thumb_images, str) else (", ".join(thumb_images) if thumb_images else "")
        image_col = image_url if image_url else (thumb_images.split(",")[0].strip() if images_col else "")
        data.append({
            "Name": name,
            "Price": price,
            "Description": desc,
            "Category": category,
            "Brand": brand,
            "Color": color,
            "Size": size,
            "Image": image_col,  # only the first image link
            "Images": images_col  # all images from #Product-Thumbnails as comma-separated string
        })

    except Exception as e:
        print(f"❌ Failed to scrape {link}: {e}")
        continue

driver.quit()

# Check if any data was scraped
if not data:
    print("⚠️ No products were scraped. Excel file will NOT be generated.")
else:
    print(f"💾 Exporting data to Excel file...")

    df = pd.DataFrame(data)

    # Save both 'Image' (first image) and 'Images' (all thumbnails)
    df_to_export = df

    output_file = "/Users/orlando/Desktop/aliexpress-scraper/files/aliexpress_products.xlsx"
    if os.path.exists(output_file):
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active
        # Add an empty row
        ws.append([])
        start_row = ws.max_row + 1
        # Don't add header again
        for r in dataframe_to_rows(df_to_export, index=False, header=False):
            ws.append(r)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        for r in dataframe_to_rows(df_to_export, index=False, header=True):
            ws.append(r)
        start_row = 2  # first data row


    # No image embedding, only image links are saved in the file

    wb.save(output_file)
    print(f"✅ Excel file saved at: {output_file}")